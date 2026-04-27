from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from gestion.models import (
    CalculoIPC,
    CalculoSalarioMinimo,
    Contrato,
    IPCHistorico,
    Local,
    SalarioMinimoHistorico,
    Tercero,
)


class CalculosAjustesTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username='calculos-ajustes', password='testpass123'
        )
        self.client.force_login(self.user)
        self.local = Local.objects.create(
            nombre_comercial_stand='Local Ajustes',
            total_area_m2=100,
        )
        self.ipc = IPCHistorico.objects.create(
            año=2025,
            valor_ipc=Decimal('9.28'),
        )
        self.salario = SalarioMinimoHistorico.objects.create(
            año=2026,
            valor_salario_minimo=Decimal('1423500'),
            variacion_porcentual=Decimal('9.54'),
        )

    def crear_contrato(self, numero):
        tercero = Tercero.objects.create(
            nit=f'903{numero}',
            razon_social=f'Cliente Ajuste {numero}',
            tipo='ARRENDATARIO',
            nombre_rep_legal='Representante',
        )
        return Contrato.objects.create(
            num_contrato=f'AJUSTE-{numero}',
            objeto_destinacion='Objeto',
            nit_concedente='800',
            rep_legal_concedente='Legal',
            fecha_firma=date(2025, 1, 1),
            fecha_inicial_contrato=date(2025, 1, 1),
            fecha_final_inicial=date(2028, 12, 31),
            modalidad_pago='Fijo',
            arrendatario=tercero,
            local=self.local,
        )

    def crear_calculo_ipc(self, contrato):
        return CalculoIPC.objects.create(
            contrato=contrato,
            año_aplicacion=2026,
            fecha_aplicacion=date(2026, 1, 1),
            ipc_historico=self.ipc,
            canon_anterior=Decimal('1000000'),
            puntos_adicionales=Decimal('0.00'),
            porcentaje_total_aplicar=Decimal('9.28'),
            valor_incremento=Decimal('92800'),
            nuevo_canon=Decimal('1092800'),
            estado='APLICADO',
        )

    def crear_calculo_salario_minimo(self, contrato):
        return CalculoSalarioMinimo.objects.create(
            contrato=contrato,
            año_aplicacion=2026,
            fecha_aplicacion=date(2026, 1, 1),
            salario_minimo_historico=self.salario,
            canon_anterior=Decimal('1000000'),
            porcentaje_salario_minimo=Decimal('9.54'),
            puntos_adicionales=Decimal('0.00'),
            porcentaje_total_aplicar=Decimal('9.54'),
            valor_incremento=Decimal('95400'),
            nuevo_canon=Decimal('1095400'),
            estado='APLICADO',
        )

    def test_lista_muestra_filtro_por_contrato_y_boton_exportar(self):
        contrato = self.crear_contrato(1)
        self.crear_calculo_ipc(contrato)

        response = self.client.get(reverse('gestion:lista_calculos_ipc'))

        self.assertContains(response, 'name="contrato"')
        self.assertContains(response, contrato.num_contrato)
        self.assertContains(response, 'Exportar Excel')
        self.assertContains(response, reverse('gestion:exportar_calculos_ajustes_excel'))

    def test_exporta_excel_de_calculos_filtrado_por_contrato(self):
        contrato_filtrado = self.crear_contrato(1)
        contrato_excluido = self.crear_contrato(2)
        self.crear_calculo_ipc(contrato_filtrado)
        self.crear_calculo_salario_minimo(contrato_excluido)

        response = self.client.get(
            reverse('gestion:exportar_calculos_ajustes_excel'),
            {'contrato': contrato_filtrado.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('calculos_ajustes', response['Content-Disposition'])

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        contenido = '\n'.join(str(valor) for row in rows for valor in row)

        self.assertIn(contrato_filtrado.num_contrato, contenido)
        self.assertNotIn(contrato_excluido.num_contrato, contenido)
