"""
Regresiones reportadas por Avenida Chile Centro Comercial y Financiero PH (revision contratos
% de ventas, septiembre 2026). Cada clase reproduce un caso real del
archivo "REVISION CONTRATOS % DE VENTAS.xlsx" con datos equivalentes.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from gestion.forms import CalculoFacturacionVentasForm, ContratoForm
from gestion.models import (
    CalculoIPC, Contrato, FinalizacionContrato, InformeVentas, IPCHistorico, Local, OtroSi,
    Tercero, TipoContrato,
)
from gestion.utils_otrosi import obtener_valores_vigentes_facturacion_ventas
from gestion.utils_ventas import contrato_reporta_ventas
from gestion.views.utils import _obtener_fecha_final_contrato, obtener_canon_vigente_con_fuente


def _crear_contrato(local, tercero, **kwargs):
    datos = dict(
        num_contrato='CTO-PRUEBA',
        tipo_contrato_cliente_proveedor='CLIENTE',
        objeto_destinacion='Objeto',
        nit_concedente='800',
        rep_legal_concedente='Legal',
        fecha_firma=date(2009, 1, 1),
        fecha_inicial_contrato=date(2009, 2, 10),
        fecha_final_inicial=date(2014, 2, 9),
        fecha_final_actualizada=date(2014, 2, 9),
        duracion_inicial_meses=60,
        modalidad_pago='Hibrido (Min Garantizado)',
        canon_minimo_garantizado=Decimal('9600000'),
        porcentaje_ventas=Decimal('12'),
        reporta_ventas=True,
        arrendatario=tercero,
        local=local,
        vigente=True,
    )
    datos.update(kwargs)
    return Contrato.objects.create(**datos)


class MinimoGarantizadoIPCSobreOtroSiTest(TestCase):
    """Caso Crepes & Waffles / Amor Perfecto: el % de ventas tomaba el minimo
    garantizado del Otro Si historico y no el del incremento anual de IPC,
    porque el Otro Si fue aprobado en la plataforma (fecha_aprobacion) despues
    de aplicar el IPC, aunque su vigencia contractual es muy anterior."""

    def setUp(self):
        self.local = Local.objects.create(nombre_comercial_stand='Local Crepes', total_area_m2=100)
        tercero = Tercero.objects.create(
            nit='860', razon_social='Crepes Prueba', tipo='ARRENDATARIO', nombre_rep_legal='Rep',
        )
        self.contrato = _crear_contrato(self.local, tercero, num_contrato='CTO-CREPES-PRUEBA')
        # OS-7: vigente desde 2023, minimo 9.500.000, aprobado en plataforma el 28/03/2026
        OtroSi.objects.create(
            contrato=self.contrato,
            numero_otrosi='OS-7',
            estado='APROBADO',
            fecha_otrosi=date(2023, 1, 1),
            effective_from=date(2023, 1, 1),
            effective_to=date(2028, 2, 9),
            nueva_modalidad_pago='Hibrido (Min Garantizado)',
            nuevo_canon_minimo_garantizado=Decimal('9500000'),
            nuevo_porcentaje_ventas=Decimal('10'),
            nueva_fecha_final_actualizada=date(2028, 2, 9),
            fecha_aprobacion=timezone.make_aware(timezone.datetime(2026, 3, 28, 21, 27)),
        )
        ipc = IPCHistorico.objects.create(año=2025, valor_ipc=Decimal('5.10'))
        CalculoIPC.objects.create(
            contrato=self.contrato,
            año_aplicacion=2026,
            fecha_aplicacion=date(2026, 1, 1),
            ipc_historico=ipc,
            canon_anterior=Decimal('10921443'),
            porcentaje_total_aplicar=Decimal('5.10'),
            valor_incremento=Decimal('556993.59'),
            nuevo_canon=Decimal('11478436.59'),
            estado='APLICADO',
        )

    def test_canon_minimo_vigente_prefiere_ipc_posterior_a_vigencia_del_otrosi(self):
        info = obtener_canon_vigente_con_fuente(
            self.contrato, date(2026, 7, 31), forzar_campo='canon_minimo_garantizado'
        )
        self.assertEqual(info['tipo'], 'ipc')
        self.assertEqual(info['valor'], Decimal('11478436.59'))

    def test_calculo_ventas_usa_minimo_ajustado_por_ipc(self):
        valores = obtener_valores_vigentes_facturacion_ventas(self.contrato, 7, 2026)
        self.assertEqual(valores['canon_minimo_garantizado'], Decimal('11478436.59'))
        self.assertIn('IPC', valores['fuente_canon_minimo_garantizado'])

    def test_otrosi_posterior_al_ipc_sigue_ganando(self):
        """Un Otro Si cuya vigencia empieza despues del IPC si debe reemplazarlo."""
        OtroSi.objects.create(
            contrato=self.contrato,
            numero_otrosi='OS-8',
            estado='APROBADO',
            fecha_otrosi=date(2026, 3, 1),
            effective_from=date(2026, 3, 1),
            nuevo_canon_minimo_garantizado=Decimal('12000000'),
            fecha_aprobacion=timezone.make_aware(timezone.datetime(2026, 3, 1, 9, 0)),
        )
        info = obtener_canon_vigente_con_fuente(
            self.contrato, date(2026, 7, 31), forzar_campo='canon_minimo_garantizado'
        )
        self.assertEqual(info['tipo'], 'otrosi')
        self.assertEqual(info['valor'], Decimal('12000000'))


@override_settings(
    ALLOWED_HOSTS=['*'],
    STORAGES={
        'default': {'BACKEND': 'django.core.files.storage.FileSystemStorage'},
        'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
    },
)
class ReportaVentasConModalidadPorOtroSiTest(TestCase):
    """Caso Perfumery / Presto: el contrato base es Fijo pero un Otro Si lo
    paso a Hibrido; el formulario no dejaba marcar 'Reporta Ventas'."""

    def setUp(self):
        self.user = User.objects.create_user(username='admin', password='x', is_staff=True)
        self.local = Local.objects.create(nombre_comercial_stand='Local Perfumery', total_area_m2=50)
        self.tercero = Tercero.objects.create(
            nit='901', razon_social='Perfumery Prueba', tipo='ARRENDATARIO', nombre_rep_legal='Rep',
        )
        self.tipo = TipoContrato.objects.create(nombre='Concesion')
        self.contrato = _crear_contrato(
            self.local, self.tercero,
            num_contrato='CTO-PERFUMERY-PRUEBA',
            tipo_contrato=self.tipo,
            fecha_inicial_contrato=date(2025, 12, 1),
            fecha_final_inicial=date(2026, 5, 31),
            fecha_final_actualizada=date(2026, 5, 31),
            duracion_inicial_meses=6,
            modalidad_pago='Fijo',
            canon_minimo_garantizado=Decimal('5000000'),
            porcentaje_ventas=None,
            reporta_ventas=False,
        )
        OtroSi.objects.create(
            contrato=self.contrato,
            numero_otrosi='OS-1',
            estado='APROBADO',
            fecha_otrosi=date(2026, 6, 1),
            effective_from=date(2026, 6, 1),
            effective_to=date(2026, 12, 31),
            nueva_modalidad_pago='Hibrido (Min Garantizado)',
            nuevo_porcentaje_ventas=Decimal('10'),
            nueva_fecha_final_actualizada=date(2026, 12, 31),
            fecha_aprobacion=timezone.make_aware(timezone.datetime(2026, 6, 18, 16, 23)),
        )

    def _datos(self, **extra):
        datos = {
            'num_contrato': self.contrato.num_contrato,
            'tipo_contrato_cliente_proveedor': 'CLIENTE',
            'tipo_contrato': self.tipo.pk,
            'arrendatario': self.tercero.pk,
            'local': self.local.pk,
            'objeto_destinacion': 'Objeto',
            'nit_concedente': '800',
            'rep_legal_concedente': 'Legal',
            'fecha_firma': '2025-11-15',
            'fecha_inicial_contrato': '2025-12-01',
            'duracion_inicial_meses': 6,
            'modalidad_pago': 'Fijo',
            'canon_minimo_garantizado': '5000000',
            'reporta_ventas': 'on',
            'dia_limite_reporte_ventas': 10,
        }
        datos.update(extra)
        return datos

    def test_permite_marcar_reporta_ventas_si_otrosi_cambio_modalidad(self):
        form = ContratoForm(data=self._datos(), instance=self.contrato, user=self.user)
        form.is_valid()
        self.assertNotIn('reporta_ventas', form.errors, form.errors)
        self.assertTrue(form.cleaned_data.get('reporta_ventas'))
        self.assertEqual(form.cleaned_data.get('dia_limite_reporte_ventas'), 10)

    def test_sigue_bloqueando_si_el_contrato_es_fijo_sin_otrosi(self):
        OtroSi.objects.all().delete()
        form = ContratoForm(data=self._datos(), instance=self.contrato, user=self.user)
        form.is_valid()
        self.assertIn('reporta_ventas', form.errors)

    def test_reporta_ventas_es_automatico_cuando_el_otrosi_trae_porcentaje(self):
        """Sin marcar la casilla del contrato base, el Otro Si ya lo hace reportar ventas."""
        self.assertFalse(self.contrato.reporta_ventas)
        self.assertTrue(contrato_reporta_ventas(self.contrato, 7, 2026))
        # Antes de la vigencia del Otro Si (mayo 2026) sigue sin reportar
        self.assertFalse(contrato_reporta_ventas(self.contrato, 5, 2026))

    def test_informe_permite_calcular_sin_tocar_el_contrato_base(self):
        informe = InformeVentas.objects.create(contrato=self.contrato, mes=7, año=2026)
        self.client.force_login(self.user)
        with patch('django.test.client.copy', lambda ctx: ctx):
            response = self.client.get(f'/informes-ventas/{informe.pk}/marcar-entregado/')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['puede_calcular'])

    def test_exportacion_muestra_reporta_ventas_si(self):
        self.client.force_login(self.user)
        with patch('gestion.views.contratos.timezone.now', return_value=timezone.make_aware(timezone.datetime(2026, 8, 20, 10, 0))):
            response = self.client.post('/exportaciones/contratos/', {'estado': '', 'tipo_contrato_cliente_proveedor': ''})
        self.assertEqual(response.status_code, 200)
        hoja = load_workbook(BytesIO(response.content))['Clientes']
        datos = dict(zip([c.value for c in hoja[1]], [c.value for c in hoja[2]]))
        self.assertEqual(datos['Reporta Ventas'], 'Sí')

    def test_vista_editar_expone_modalidad_vigente_para_el_javascript(self):
        self.client.force_login(self.user)
        with patch('django.test.client.copy', lambda ctx: ctx):
            response = self.client.get(f'/contratos/{self.contrato.pk}/editar/')
        self.assertEqual(response.status_code, 200)
        self.assertIn('modalidad-pago-vigente', response.content.decode())
        self.assertIn('Hibrido (Min Garantizado)', response.content.decode())


@override_settings(
    ALLOWED_HOSTS=['*'],
    STORAGES={
        'default': {'BACKEND': 'django.core.files.storage.FileSystemStorage'},
        'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
    },
)
class ContratoTerminadoEnReportesTest(TestCase):
    """Caso Servientrega: el contrato se dio por terminado en la plataforma
    pero seguia saliendo como 'Vigente' en la exportacion y en la lista de
    informes de ventas."""

    def setUp(self):
        self.user = User.objects.create_user(username='admin', password='x', is_staff=True)
        self.client.force_login(self.user)
        self.local = Local.objects.create(nombre_comercial_stand='Local Servientrega', total_area_m2=20)
        tercero = Tercero.objects.create(
            nit='890', razon_social='Servientrega Prueba', tipo='ARRENDATARIO', nombre_rep_legal='Rep',
        )
        self.contrato = _crear_contrato(
            self.local, tercero,
            num_contrato='CTO-SERVIENTREGA-PRUEBA',
            fecha_firma=date(2025, 9, 15),
            fecha_inicial_contrato=date(2025, 10, 1),
            fecha_final_inicial=date(2026, 9, 30),
            fecha_final_actualizada=date(2026, 9, 30),
            duracion_inicial_meses=12,
            modalidad_pago='Variable Puro',
            canon_minimo_garantizado=Decimal('0'),
            porcentaje_ventas=Decimal('8'),
        )
        FinalizacionContrato.objects.create(
            contrato=self.contrato,
            motivo='MUTUO_ACUERDO',
            fecha_finalizacion=date(2026, 6, 30),
            acuerdo_partes=True,
            registrado_por=self.user,
        )
        self.contrato.finalizado = True
        self.contrato.save(update_fields=['finalizado'])

    def test_fecha_final_respeta_la_terminacion(self):
        self.assertEqual(_obtener_fecha_final_contrato(self.contrato, date(2026, 8, 31)), date(2026, 6, 30))

    def test_exportacion_marca_terminado(self):
        with patch('gestion.views.contratos.timezone.now', return_value=timezone.make_aware(timezone.datetime(2026, 8, 20, 10, 0))):
            response = self.client.post('/exportaciones/contratos/', {'estado': '', 'tipo_contrato_cliente_proveedor': ''})
        self.assertEqual(response.status_code, 200, getattr(response, 'url', ''))
        hoja = load_workbook(BytesIO(response.content))['Clientes']
        datos = dict(zip([c.value for c in hoja[1]], [c.value for c in hoja[2]]))
        self.assertEqual(datos['Estado'], 'Terminado')
        self.assertEqual(datos['Fecha Final Actualizada'].date() if hasattr(datos['Fecha Final Actualizada'], 'date') else datos['Fecha Final Actualizada'], date(2026, 6, 30))

    def test_exportacion_vigentes_excluye_terminado(self):
        with patch('gestion.views.contratos.timezone.now', return_value=timezone.make_aware(timezone.datetime(2026, 8, 20, 10, 0))):
            response = self.client.post('/exportaciones/contratos/', {'estado': 'vigentes', 'tipo_contrato_cliente_proveedor': ''})
        self.assertEqual(response.status_code, 302)  # sin resultados -> redirect con warning

    def test_lista_informes_ventas_excluye_terminado_tras_la_fecha(self):
        with patch('django.test.client.copy', lambda ctx: ctx):
            response = self.client.get('/informes-ventas/', {'mes': '8', 'año': '2026'})
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(self.contrato, [i['contrato'] for i in response.context['contratos_info']])

    def test_lista_informes_ventas_incluye_antes_de_la_fecha(self):
        with patch('django.test.client.copy', lambda ctx: ctx):
            response = self.client.get('/informes-ventas/', {'mes': '5', 'año': '2026'})
        self.assertEqual(response.status_code, 200)
        self.assertIn(self.contrato, [i['contrato'] for i in response.context['contratos_info']])


class MensajeFueraDeVigenciaTest(TestCase):
    """Caso La Bonga / Carbon 100 (enero): el contrato base termino en 2014 y el
    Otro Si vigente arranca en febrero 2026; enero queda por fuera de la
    vigencia registrada. El mensaje debe decirlo, no hablar de la modalidad."""

    def setUp(self):
        self.local = Local.objects.create(nombre_comercial_stand='Local Bonga', total_area_m2=100)
        tercero = Tercero.objects.create(
            nit='900', razon_social='Renacimiento Prueba', tipo='ARRENDATARIO', nombre_rep_legal='Rep',
        )
        self.contrato = _crear_contrato(
            self.local, tercero,
            num_contrato='2025-101-PRUEBA',
            fecha_inicial_contrato=date(2009, 2, 27),
            fecha_final_inicial=date(2014, 2, 26),
            fecha_final_actualizada=date(2014, 2, 26),
        )
        OtroSi.objects.create(
            contrato=self.contrato,
            numero_otrosi='OS-7',
            estado='APROBADO',
            fecha_otrosi=date(2026, 2, 27),
            effective_from=date(2026, 2, 27),
            effective_to=date(2028, 2, 26),
            nueva_modalidad_pago='Hibrido (Min Garantizado)',
            nuevo_canon_minimo_garantizado=Decimal('8943333'),
            nuevo_porcentaje_ventas=Decimal('10'),
            nueva_fecha_final_actualizada=date(2028, 2, 26),
        )

    def _form(self, mes):
        return CalculoFacturacionVentasForm(data={
            'contrato': self.contrato.id,
            'mes': str(mes),
            'año': 2026,
            'ventas_totales': '50000000',
            'devoluciones': '0',
            'observaciones': '',
        })

    def test_febrero_es_valido(self):
        form = self._form(2)
        self.assertTrue(form.is_valid(), form.errors)

    def test_enero_explica_que_esta_fuera_de_vigencia(self):
        form = self._form(1)
        self.assertFalse(form.is_valid())
        mensaje = ' '.join(form.non_field_errors())
        self.assertIn('no estaba vigente', mensaje)
        self.assertIn('26/02/2014', mensaje)
        self.assertNotIn('modalidad', mensaje)
