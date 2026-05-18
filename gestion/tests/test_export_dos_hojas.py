from django.test import TestCase
from gestion.utils_otrosi import get_valores_polizas_vigentes, _CAMPOS_POLIZAS_EXPORTACION


class CamposPolizasExportacionTest(TestCase):

    def test_campos_proveedor_rce_en_lista(self):
        campos = [campo_nuevo for campo_nuevo, _ in _CAMPOS_POLIZAS_EXPORTACION]
        self.assertIn('nuevo_rce_cobertura_danos_materiales', campos)
        self.assertIn('nuevo_rce_cobertura_perjuicios_patrimoniales', campos)

    def test_campos_proveedor_cumplimiento_en_lista(self):
        campos = [campo_nuevo for campo_nuevo, _ in _CAMPOS_POLIZAS_EXPORTACION]
        self.assertIn('nuevo_cumplimiento_amparo_cumplimiento_contrato', campos)
        self.assertIn('nuevo_cumplimiento_amparo_sanciones_incumplimiento', campos)

    def test_total_campos_polizas(self):
        # 47 originales + 12 rce_cobertura + 11 cumplimiento_amparo = 70
        self.assertEqual(len(_CAMPOS_POLIZAS_EXPORTACION), 70)


from io import BytesIO
from openpyxl import load_workbook
from gestion.services.exportes import (
    ColumnaExportacion,
    ExportacionVaciaError,
    generar_excel_multi_hoja,
)


COLUMNAS_SIMPLES = [
    ColumnaExportacion('Columna A', ancho=15),
    ColumnaExportacion('Columna B', ancho=15, es_numerica=True),
]


class GenerarExcelMultiHojaTest(TestCase):

    def _cargar(self, resultado: bytes):
        return load_workbook(BytesIO(resultado))

    def test_crea_dos_hojas(self):
        resultado = generar_excel_multi_hoja([
            ('Clientes', COLUMNAS_SIMPLES, [('Acme', 100)]),
            ('Proveedores', COLUMNAS_SIMPLES, [('Beta', 200)]),
        ])
        wb = self._cargar(resultado)
        self.assertEqual(wb.sheetnames, ['Clientes', 'Proveedores'])

    def test_hoja_vacia_muestra_mensaje(self):
        resultado = generar_excel_multi_hoja([
            ('Clientes', COLUMNAS_SIMPLES, [('Acme', 100)]),
            ('Proveedores', COLUMNAS_SIMPLES, []),
        ])
        wb = self._cargar(resultado)
        hoja = wb['Proveedores']
        # Fila 1 = encabezados, fila 2 = mensaje
        self.assertEqual(hoja.cell(row=2, column=1).value, 'Sin contratos registrados')

    def test_error_si_todas_vacias(self):
        with self.assertRaises(ExportacionVaciaError):
            generar_excel_multi_hoja([
                ('Clientes', COLUMNAS_SIMPLES, []),
                ('Proveedores', COLUMNAS_SIMPLES, []),
            ])

    def test_datos_escritos_correctamente(self):
        resultado = generar_excel_multi_hoja([
            ('Clientes', COLUMNAS_SIMPLES, [('Texto', 42)]),
            ('Proveedores', COLUMNAS_SIMPLES, []),
        ])
        wb = self._cargar(resultado)
        hoja = wb['Clientes']
        self.assertEqual(hoja.cell(row=1, column=1).value, 'Columna A')
        self.assertEqual(hoja.cell(row=2, column=1).value, 'Texto')
        self.assertEqual(hoja.cell(row=2, column=2).value, 42)

    def test_error_si_registro_tiene_columnas_incorrectas(self):
        with self.assertRaises(ValueError):
            generar_excel_multi_hoja([
                ('Clientes', COLUMNAS_SIMPLES, [('solo uno',)]),  # esperaba 2 valores
            ])


from django.test import Client
from django.urls import reverse
from django.contrib.auth.models import User
from gestion.models import (
    Contrato, Local, Tercero,
)
from datetime import date


class ExportarContratosDocsHojasTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='tester', password='pass')
        self.client.force_login(self.user)

        self.local = Local.objects.create(
            nombre_comercial_stand='Local Test',
            ubicacion='Piso 1',
            total_area_m2='50.00',
        )
        self.tercero = Tercero.objects.create(
            razon_social='Empresa Test',
            nit='900000001',
            nombre_rep_legal='Rep Legal Test',
        )
        Contrato.objects.create(
            num_contrato='TEST-CLIENTE-001',
            tipo_contrato_cliente_proveedor='CLIENTE',
            local=self.local,
            arrendatario=self.tercero,
            fecha_firma=date(2024, 1, 1),
            fecha_inicial_contrato=date(2024, 1, 1),
            fecha_final_inicial=date(2025, 12, 31),
            objeto_destinacion='Objeto test',
            nit_concedente='800000001',
            rep_legal_concedente='Rep Legal Concedente',
        )
        Contrato.objects.create(
            num_contrato='TEST-PROVEEDOR-001',
            tipo_contrato_cliente_proveedor='PROVEEDOR',
            local=self.local,
            proveedor=self.tercero,
            fecha_firma=date(2024, 1, 1),
            fecha_inicial_contrato=date(2024, 1, 1),
            fecha_final_inicial=date(2025, 12, 31),
            objeto_destinacion='Objeto test proveedor',
            nit_concedente='800000001',
            rep_legal_concedente='Rep Legal Concedente',
        )

    def _exportar(self):
        response = self.client.post(reverse('gestion:exportar_contratos'), {
            'exportar': '1',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        return load_workbook(BytesIO(response.content))

    def test_export_genera_dos_hojas(self):
        wb = self._exportar()
        self.assertIn('Clientes', wb.sheetnames)
        self.assertIn('Proveedores', wb.sheetnames)

    def test_hoja_clientes_contiene_contrato_cliente(self):
        wb = self._exportar()
        valores = [wb['Clientes'].cell(row=2, column=1).value]
        self.assertIn('TEST-CLIENTE-001', valores)

    def test_hoja_proveedores_contiene_contrato_proveedor(self):
        wb = self._exportar()
        valores = [wb['Proveedores'].cell(row=2, column=1).value]
        self.assertIn('TEST-PROVEEDOR-001', valores)

    def test_hoja_vacia_muestra_sin_contratos(self):
        Contrato.objects.filter(
            tipo_contrato_cliente_proveedor='PROVEEDOR'
        ).delete()
        wb = self._exportar()
        self.assertEqual(
            wb['Proveedores'].cell(row=2, column=1).value,
            'Sin contratos registrados',
        )
