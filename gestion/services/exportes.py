"""
Servicios para la construcción de archivos de exportación.
"""

from dataclasses import dataclass
from io import BytesIO
from typing import Sequence

from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def formatear_numero_con_puntos(valor):
    """Formatea un número con punto como separador de miles y sin decimales"""
    if valor is None:
        return None
    try:
        valor_int = int(round(float(valor)))
        # Formatear con punto como separador de miles
        valor_str = f"{valor_int:,}".replace(',', '.')
        return valor_str
    except (ValueError, TypeError):
        return None

PALETA_AVENIDA = {
    'verde': 'FF8BC34A',
    'naranja': 'FFFF9800',
    'cian': 'FF00BCD4',
    'magenta': 'FFE91E63',
    'oscuro': 'FF2C3E50',
    'claro': 'FFF8F9FA',
    'blanco': 'FFFFFFFF',
    'gris_claro': 'FFE0E0E0',
    'resalte': 'FFFFF3E0',
}

FORMATO_NUMERICO = '#.##0'


def limpiar_nombre_hoja_excel(nombre: str) -> str:
    """
    Limpia el nombre de una hoja de Excel eliminando caracteres no permitidos.
    Excel no permite: / \ ? * [ ]
    También limita la longitud a 31 caracteres.
    """
    caracteres_invalidos = ['/', '\\', '?', '*', '[', ']']
    nombre_limpio = nombre
    for char in caracteres_invalidos:
        nombre_limpio = nombre_limpio.replace(char, ' ')
    # Limpiar espacios múltiples
    nombre_limpio = ' '.join(nombre_limpio.split())
    # Limitar longitud a 31 caracteres (límite de Excel)
    if len(nombre_limpio) > 31:
        nombre_limpio = nombre_limpio[:31].rstrip()
    return nombre_limpio


@dataclass(frozen=True)
class ColumnaExportacion:
    titulo: str
    ancho: int | None = None
    es_numerica: bool = False
    alineacion: str = 'left'


class FormateadorExcelCorporativo:
    """
    Aplica formato corporativo consistente a las hojas Excel exportadas.
    """

    def __init__(self, columnas: Sequence[ColumnaExportacion]) -> None:
        self.columnas = columnas
        self._bordes_celda = Border(
            left=Side(style='thin', color=PALETA_AVENIDA['gris_claro']),
            right=Side(style='thin', color=PALETA_AVENIDA['gris_claro']),
            top=Side(style='thin', color=PALETA_AVENIDA['gris_claro']),
            bottom=Side(style='thin', color=PALETA_AVENIDA['gris_claro']),
        )

    def aplicar(self, hoja, total_filas_datos: int) -> None:
        hoja.sheet_view.showGridLines = False
        hoja.freeze_panes = 'A2'
        hoja.row_dimensions[1].height = 24

        self._formatear_encabezados(hoja)
        if total_filas_datos > 1:
            self._formatear_cuerpo(hoja, total_filas_datos)
        self._ajustar_anchos_columnas(hoja, total_filas_datos)

    def _formatear_encabezados(self, hoja) -> None:
        encabezado_font = Font(name='Arial', size=9, color=PALETA_AVENIDA['blanco'], bold=True)
        encabezado_fill = PatternFill(fill_type='solid', start_color=PALETA_AVENIDA['oscuro'], end_color=PALETA_AVENIDA['oscuro'])

        for indice, celda in enumerate(hoja[1], start=1):
            col_config = self.columnas[indice - 1]
            celda.value = str(col_config.titulo).strip()
            celda.font = encabezado_font
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = self._bordes_celda

    def _formatear_cuerpo(self, hoja, ultima_fila_datos: int) -> None:
        for fila in range(2, ultima_fila_datos + 1):
            es_par = fila % 2 == 0
            relleno = PatternFill(
                fill_type='solid',
                start_color=PALETA_AVENIDA['claro'] if es_par else PALETA_AVENIDA['blanco'],
                end_color=PALETA_AVENIDA['claro'] if es_par else PALETA_AVENIDA['blanco'],
            )

            for columna in range(1, len(self.columnas) + 1):
                celda = hoja.cell(row=fila, column=columna)
                configuracion = self.columnas[columna - 1]
                celda.fill = relleno
                celda.border = self._bordes_celda
                celda.font = Font(name='Arial', size=9)

                # Solo aplicar formato numérico si el valor es realmente un número (int o float)
                # No aplicar a strings aunque la columna esté marcada como numérica
                if configuracion.es_numerica and celda.value is not None and celda.value != 'N/A':
                    valor_original = celda.value
                    # Verificar que sea realmente un número, no un string (como porcentajes o texto)
                    if isinstance(valor_original, (int, float)) and not isinstance(valor_original, str):
                        valor_entero = int(round(float(valor_original)))
                        # Guardar como número para que Excel pueda hacer cálculos
                        celda.value = valor_entero
                        # Aplicar formato numérico estándar
                        # El formato #,##0 usa separador de miles según configuración regional de Excel
                        # Si Excel está configurado en español, usará punto como separador
                        celda.number_format = '#,##0'
                        celda.alignment = Alignment(horizontal='right', vertical='center')
                    elif isinstance(valor_original, str) and valor_original.replace('.', '').replace(',', '').isdigit():
                        # Si viene como string numérico puro (sin caracteres especiales), convertir a int
                        celda.value = int(float(valor_original.replace(',', '.')))
                        celda.number_format = '#,##0'
                        celda.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        # Si es string con texto (como "20%" o "Sí"), mantener como texto
                        celda.alignment = self._resolver_alineacion(configuracion.alineacion)
                else:
                    celda.alignment = self._resolver_alineacion(configuracion.alineacion)

    def _ajustar_anchos_columnas(self, hoja, ultima_fila_datos: int) -> None:
        for indice, configuracion in enumerate(self.columnas, start=1):
            if configuracion.ancho is not None:
                ancho = configuracion.ancho
            else:
                ancho = self._calcular_ancho_columna(hoja, indice, ultima_fila_datos)
            columna_excel = get_column_letter(indice)
            hoja.column_dimensions[columna_excel].width = max(ancho, 12)

    def _calcular_ancho_columna(self, hoja, indice_columna: int, ultima_fila: int) -> int:

        def _representacion(valor) -> str:
            if valor is None:
                return ''
            if isinstance(valor, (int, float)):
                return self._formatear_numero_para_ancho(valor)
            return str(valor)

        max_largo = 0
        for fila in range(1, ultima_fila + 1):
            valor = hoja.cell(row=fila, column=indice_columna).value
            longitud = len(_representacion(valor))
            max_largo = max(max_largo, longitud)
        return max_largo + 2

    @staticmethod
    def _resolver_alineacion(valor: str) -> Alignment:
        alineacion = valor.lower()
        if alineacion == 'center':
            return Alignment(horizontal='center', vertical='center')
        if alineacion == 'right':
            return Alignment(horizontal='right', vertical='center')
        return Alignment(horizontal='left', vertical='center')

    @staticmethod
    def _formatear_numero_para_ancho(valor: int | float) -> str:
        if isinstance(valor, float):
            valor = int(round(valor))
        return f"{valor:,}".replace(',', '.')


class ExportacionVaciaError(Exception):
    """Se lanza cuando no hay información para exportar."""


def generar_excel_corporativo(
    nombre_hoja: str,
    columnas: Sequence[ColumnaExportacion],
    registros: Sequence[Sequence],
) -> bytes:
    """
    Construye un archivo XLSX aplicando el formato corporativo estándar.
    """
    if not registros:
        raise ExportacionVaciaError('No hay información disponible para exportar.')

    formateador = FormateadorExcelCorporativo(columnas)

    libro = Workbook()
    hoja = libro.active
    hoja.title = limpiar_nombre_hoja_excel(nombre_hoja)

    hoja.append([columna.titulo for columna in columnas])
    num_columnas = len(columnas)
    
    for registro in registros:
        registro_procesado = []
        num_valores = len(registro)
        
        # Validar que el registro tenga el mismo número de valores que columnas
        if num_valores != num_columnas:
            raise ValueError(
                f'El registro tiene {num_valores} valores pero se esperaban {num_columnas} columnas. '
                f'Registro: {registro[:5]}...'
            )
        
        for indice, valor in enumerate(registro):
            if indice >= num_columnas:
                break
            col_config = columnas[indice]
            # Si es numérico y el valor es None o vacío, convertir a "N/A"
            if col_config.es_numerica:
                # Si el valor es un string (como porcentajes o texto), no tratarlo como numérico
                if isinstance(valor, str):
                    registro_procesado.append(valor)
                elif valor is None:
                    registro_procesado.append('N/A')
                else:
                    try:
                        # Convertir a entero (sin decimales)
                        valor_numerico = float(valor)
                        valor_entero = int(round(valor_numerico))
                        # Guardar como número (no como texto) para que Excel pueda aplicar formato
                        registro_procesado.append(valor_entero)
                    except (ValueError, TypeError):
                        registro_procesado.append('N/A')
            else:
                # Para campos no numéricos, None o vacío se convierte a "N/A"
                if valor is None or valor == '' or (isinstance(valor, str) and valor.strip() == ''):
                    registro_procesado.append('N/A')
                else:
                    registro_procesado.append(valor)
        hoja.append(registro_procesado)

    ultima_fila_datos = len(registros) + 1
    formateador.aplicar(hoja, ultima_fila_datos)

    hoja.append([])
    ahora_local = timezone.localtime(timezone.now())
    hoja.append(
        [
            f'Generado el {ahora_local.strftime("%Y-%m-%d %H:%M:%S")}',
        ]
    )
    fila_informativa = hoja.max_row
    hoja.merge_cells(start_row=fila_informativa, start_column=1, end_row=fila_informativa, end_column=len(columnas))
    celda_info = hoja.cell(row=fila_informativa, column=1)
    celda_info.alignment = Alignment(horizontal='right', vertical='center')
    celda_info.font = Font(name='Arial', size=9, italic=True, color=PALETA_AVENIDA['oscuro'])

    buffer = BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel_informes_ventas(informes_queryset=None):
    """
    Genera un archivo Excel detallado con los informes de ventas especificados.
    Incluye información completa del informe, contrato, arrendatario, local y cálculos asociados.
    
    Args:
        informes_queryset: QuerySet opcional de InformeVentas. Si no se proporciona, 
                          se obtienen todos los informes.
    """
    from gestion.models import InformeVentas, CalculoFacturacionVentas
    
    # Obtener los informes con relaciones optimizadas
    if informes_queryset is None:
        informes = InformeVentas.objects.select_related(
            'contrato', 'contrato__arrendatario', 'contrato__local', 'contrato__tipo_contrato'
        ).prefetch_related('calculos_facturacion').order_by('-año', '-mes', 'contrato__num_contrato')
    else:
        # Si se proporciona un queryset, asegurar que tenga las relaciones optimizadas
        informes = informes_queryset.select_related(
            'contrato', 'contrato__arrendatario', 'contrato__local', 'contrato__tipo_contrato'
        ).prefetch_related('calculos_facturacion').order_by('-año', '-mes', 'contrato__num_contrato')
    
    def formatear_moneda_excel(valor):
        """Formatea un valor decimal como número entero para Excel"""
        if valor is None:
            return 0
        return int(round(float(valor)))
    
    # Definir columnas detalladas
    columnas = [
        # Información del Informe
        ColumnaExportacion('Número Contrato', ancho=15),
        ColumnaExportacion('Arrendatario', ancho=30),
        ColumnaExportacion('NIT Arrendatario', ancho=15),
        ColumnaExportacion('Local', ancho=20),
        ColumnaExportacion('Tipo Contrato', ancho=20),
        ColumnaExportacion('Mes', ancho=12),
        ColumnaExportacion('Año', ancho=8, es_numerica=True),
        ColumnaExportacion('Estado', ancho=12),
        ColumnaExportacion('Fecha Entrega', ancho=12),
        ColumnaExportacion('Días Vencido', ancho=12, es_numerica=True),
        ColumnaExportacion('Observaciones Informe', ancho=30),
        ColumnaExportacion('Registrado Por', ancho=20),
        ColumnaExportacion('Fecha Registro', ancho=18),
        
        # Información del Contrato
        ColumnaExportacion('Modalidad Pago', ancho=20),
        ColumnaExportacion('Canon Fijo', ancho=15, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Canon Mínimo Garantizado', ancho=25, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Porcentaje Ventas (%)', ancho=18, es_numerica=False, alineacion='right'),
        ColumnaExportacion('Día Límite Reporte', ancho=18, es_numerica=True),
        
        # Información del Cálculo (si existe)
        ColumnaExportacion('Tiene Cálculo', ancho=12),
        ColumnaExportacion('Ventas Totales', ancho=18, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Devoluciones', ancho=15, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Base Neta', ancho=15, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Porcentaje Vigente (%)', ancho=20, es_numerica=False, alineacion='right'),
        ColumnaExportacion('Valor Calculado (Base × %)', ancho=25, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Canon Mínimo Vigente', ancho=22, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Excedente sobre Mínimo', ancho=22, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Aplica Variable', ancho=15),
        ColumnaExportacion('Valor a Facturar Variable', ancho=25, es_numerica=True, alineacion='right'),
        ColumnaExportacion('Otro Sí Referencia', ancho=20),
        ColumnaExportacion('Calculado Por', ancho=20),
        ColumnaExportacion('Fecha Cálculo', ancho=18),
        ColumnaExportacion('Observaciones Cálculo', ancho=30),
    ]
    
    registros = []
    
    for informe in informes:
        # Obtener el último cálculo asociado (si existe)
        calculo = informe.calculos_facturacion.order_by('-fecha_calculo').first()
        
        # Información básica del informe
        registro = [
            informe.contrato.num_contrato,
            informe.contrato.arrendatario.razon_social,
            informe.contrato.arrendatario.nit,
            informe.contrato.local.nombre_comercial_stand,
            informe.contrato.tipo_contrato.nombre if informe.contrato.tipo_contrato else 'N/A',
            informe.get_mes_display(),
            informe.año,
            informe.get_estado_display(),
            informe.fecha_entrega.strftime('%d/%m/%Y') if informe.fecha_entrega else 'N/A',
            informe.dias_vencido() if informe.esta_vencido() else 0,
            informe.observaciones or 'N/A',
            informe.registrado_por or 'N/A',
                timezone.localtime(informe.fecha_registro).strftime('%d/%m/%Y %H:%M') if informe.fecha_registro else 'N/A',
            
            # Información del contrato
            informe.contrato.modalidad_pago or 'N/A',
            formatear_moneda_excel(informe.contrato.valor_canon_fijo),
            formatear_moneda_excel(informe.contrato.canon_minimo_garantizado),
            f'{informe.contrato.porcentaje_ventas}%' if informe.contrato.porcentaje_ventas else 'N/A',
            informe.contrato.dia_limite_reporte_ventas or 'N/A',
        ]
        
        # Información del cálculo (si existe)
        # Total de columnas de cálculo: 14
        if calculo:
            registro.extend([
                'Sí',  # 1. Tiene Cálculo
                formatear_moneda_excel(calculo.ventas_totales) if calculo.ventas_totales is not None else 'N/A',  # 2. Ventas Totales
                formatear_moneda_excel(calculo.devoluciones) if calculo.devoluciones is not None else 'N/A',  # 3. Devoluciones
                formatear_moneda_excel(calculo.base_neta) if calculo.base_neta is not None else 'N/A',  # 4. Base Neta
                f'{calculo.porcentaje_ventas_vigente}%' if calculo.porcentaje_ventas_vigente is not None else 'N/A',  # 5. Porcentaje Vigente (%)
                formatear_moneda_excel(calculo.valor_calculado_porcentaje) if calculo.valor_calculado_porcentaje is not None else 'N/A',  # 6. Valor Calculado (Base × %)
                formatear_moneda_excel(calculo.canon_minimo_garantizado_vigente) if calculo.canon_minimo_garantizado_vigente is not None else 'N/A',  # 7. Canon Mínimo Vigente
                formatear_moneda_excel(calculo.excedente_sobre_minimo) if calculo.excedente_sobre_minimo is not None else 'N/A',  # 8. Excedente sobre Mínimo
                'Sí' if calculo.aplica_variable else 'No',  # 9. Aplica Variable
                formatear_moneda_excel(calculo.valor_a_facturar_variable) if calculo.valor_a_facturar_variable is not None else 'N/A',  # 10. Valor a Facturar Variable
                calculo.otrosi_referencia.numero_otrosi if calculo.otrosi_referencia else 'N/A',  # 11. Otro Sí Referencia
                calculo.calculado_por or 'N/A',  # 12. Calculado Por
                timezone.localtime(calculo.fecha_calculo).strftime('%d/%m/%Y %H:%M') if calculo.fecha_calculo else 'N/A',  # 13. Fecha Cálculo (zona horaria Colombia)
                calculo.observaciones or 'N/A',  # 14. Observaciones Cálculo
            ])
        else:
            # Cuando no hay cálculo, agregar exactamente 14 valores 'N/A'
            registro.extend([
                'No',  # 1. Tiene Cálculo
                'N/A',  # 2. Ventas Totales
                'N/A',  # 3. Devoluciones
                'N/A',  # 4. Base Neta
                'N/A',  # 5. Porcentaje Vigente (%)
                'N/A',  # 6. Valor Calculado (Base × %)
                'N/A',  # 7. Canon Mínimo Vigente
                'N/A',  # 8. Excedente sobre Mínimo
                'N/A',  # 9. Aplica Variable
                'N/A',  # 10. Valor a Facturar Variable
                'N/A',  # 11. Otro Sí Referencia
                'N/A',  # 12. Calculado Por
                'N/A',  # 13. Fecha Cálculo
                'N/A',  # 14. Observaciones Cálculo
            ])
        
        registros.append(registro)
    
    ahora_local = timezone.localtime(timezone.now())
    nombre_hoja = f'Informes Ventas {ahora_local.strftime("%Y%m%d")}'
    
    return generar_excel_corporativo(
        nombre_hoja=nombre_hoja,
        columnas=columnas,
        registros=registros,
    )


def generar_excel_calculo_facturacion(calculo):
    """
    Genera un archivo Excel con el desglose completo del cálculo de facturación por ventas.
    """
    from gestion.models import CalculoFacturacionVentas
    
    def formatear_moneda_excel(valor):
        """Formatea un valor decimal como número entero para Excel"""
        if valor is None:
            return 0
        return int(round(float(valor)))
    
    columnas = [
        ColumnaExportacion('Concepto', ancho=40),
        ColumnaExportacion('Valor', ancho=25, es_numerica=True, alineacion='right'),
    ]
    
    registros = [
        ('Ventas Totales', formatear_moneda_excel(calculo.ventas_totales)),
        ('Devoluciones', formatear_moneda_excel(calculo.devoluciones)),
        ('Base Neta', formatear_moneda_excel(calculo.base_neta)),
        ('Porcentaje de Ventas Vigente', f'{calculo.porcentaje_ventas_vigente}%'),
        ('Valor Calculado (Base × %)', formatear_moneda_excel(calculo.valor_calculado_porcentaje)),
    ]
    
    if calculo.modalidad_contrato == 'HIBRIDO_MIN_GARANTIZADO':
        registros.append(('Canon Mínimo Garantizado Vigente', formatear_moneda_excel(calculo.canon_minimo_garantizado_vigente)))
        if calculo.excedente_sobre_minimo:
            registros.append(('Excedente sobre Mínimo', formatear_moneda_excel(calculo.excedente_sobre_minimo)))
        # Este campo es texto, no numérico, así que lo agregamos como texto
        registros.append(('¿Aplica Variable?', 'Sí' if calculo.aplica_variable else 'No (Solo Mínimo Garantizado)'))
    
    registros.append(('VALOR A FACTURAR VARIABLE', formatear_moneda_excel(calculo.valor_a_facturar_variable)))
    
    # Limpiar nombre de hoja: Excel no permite caracteres especiales como /, \, ?, *, [, ]
    nombre_mes = calculo.get_mes_display()
    nombre_hoja = f'Cálculo {nombre_mes} {calculo.año}'
    
    return generar_excel_corporativo(
        nombre_hoja=nombre_hoja,
        columnas=columnas,
        registros=registros,
    )


def generar_pdf_calculo_facturacion(calculo, configuracion_empresa):
    """
    Genera un PDF profesional con el desglose completo del cálculo de facturación por ventas.
    Incluye información de la empresa y formato corporativo.
    """
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.colors import HexColor
    import os
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Colores corporativos
    verde = HexColor('#8BC34A')
    naranja = HexColor('#FF9800')
    cian = HexColor('#00BCD4')
    magenta = HexColor('#E91E63')
    oscuro = HexColor('#2C3E50')
    
    # Crear estilos personalizados
    titulo_estilo = ParagraphStyle(
        'TituloCustom',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=oscuro,
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
    )
    
    subtitulo_estilo = ParagraphStyle(
        'SubtituloCustom',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=oscuro,
        spaceAfter=10,
        fontName='Helvetica-Bold',
    )
    
    texto_estilo = ParagraphStyle(
        'TextoCustom',
        parent=styles['Normal'],
        fontSize=10,
        textColor=oscuro,
        spaceAfter=6,
    )
    
    # Contenido del PDF
    story = []
    
    # Encabezado con información de la empresa
    header_data = [
        [configuracion_empresa.nombre_empresa if configuracion_empresa else 'Centro Comercial Avenida de Chile - PH'],
        [f'NIT: {configuracion_empresa.nit_empresa if configuracion_empresa else "860.509.249-3"}'],
    ]
    
    if configuracion_empresa:
        if configuracion_empresa.direccion:
            header_data.append([configuracion_empresa.direccion])
        if configuracion_empresa.telefono:
            header_data.append([f'Teléfono: {configuracion_empresa.telefono}'])
        if configuracion_empresa.email:
            header_data.append([f'Email: {configuracion_empresa.email}'])
    
    header_table = Table(header_data, colWidths=[7*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (0, 0), 16),
        ('TEXTCOLOR', (0, 0), (0, 0), oscuro),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TEXTCOLOR', (0, 1), (-1, -1), oscuro),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(header_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Título del documento
    story.append(Paragraph('CÁLCULO DE FACTURACIÓN POR VENTAS', titulo_estilo))
    story.append(Spacer(1, 0.2*inch))
    
    # Información del contrato
    # Convertir fecha_calculo a zona horaria de Colombia
    fecha_calculo_local = timezone.localtime(calculo.fecha_calculo) if calculo.fecha_calculo else None
    fecha_calculo_str = fecha_calculo_local.strftime('%d/%m/%Y %H:%M') if fecha_calculo_local else 'N/A'
    
    info_contrato = [
        ['Contrato:', calculo.contrato.num_contrato],
        ['Arrendatario:', calculo.contrato.arrendatario.razon_social],
        ['Local:', calculo.contrato.local.nombre_comercial_stand],
        ['Mes/Año:', f'{calculo.get_mes_display()}/{calculo.año}'],
        ['Modalidad:', calculo.get_modalidad_contrato_display()],
        ['Fecha de Cálculo:', fecha_calculo_str],
    ]
    
    info_table = Table(info_contrato, colWidths=[2*inch, 5*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), oscuro),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), HexColor('#F8F9FA')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Desglose del cálculo
    story.append(Paragraph('DESGLOSE DEL CÁLCULO', subtitulo_estilo))
    
    def formatear_moneda(valor):
        """Formatea un valor decimal como moneda con puntos como separador de miles"""
        if valor is None:
            return '$0,00'
        valor_str = f'{valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return f'${valor_str}'
    
    desglose_data = [
        ['Concepto', 'Valor'],
        ['Ventas Totales', formatear_moneda(calculo.ventas_totales)],
        ['Devoluciones', formatear_moneda(calculo.devoluciones)],
        ['Base Neta', formatear_moneda(calculo.base_neta)],
        ['Porcentaje de Ventas Vigente', f'{calculo.porcentaje_ventas_vigente}%'],
        ['Valor Calculado (Base × %)', formatear_moneda(calculo.valor_calculado_porcentaje)],
    ]
    
    # Índices de filas que deben estar en negrita
    filas_negrita = [3, 5]  # Base Neta y Valor Calculado
    
    if calculo.modalidad_contrato == 'HIBRIDO_MIN_GARANTIZADO':
        desglose_data.append(['Canon Mínimo Garantizado Vigente', formatear_moneda(calculo.canon_minimo_garantizado_vigente)])
        if calculo.excedente_sobre_minimo:
            desglose_data.append(['Excedente sobre Mínimo', formatear_moneda(calculo.excedente_sobre_minimo)])
            filas_negrita.append(len(desglose_data) - 1)
        desglose_data.append(['¿Aplica Variable?', 'Sí' if calculo.aplica_variable else 'No (Solo Mínimo Garantizado)'])
    
    # Agregar fila final de valor a facturar
    fila_final = len(desglose_data)
    desglose_data.append(['VALOR A FACTURAR VARIABLE', formatear_moneda(calculo.valor_a_facturar_variable)])
    
    desglose_table = Table(desglose_data, colWidths=[4.5*inch, 2.5*inch])
    
    # Crear estilo base
    estilo_tabla = [
        ('BACKGROUND', (0, 0), (-1, 0), oscuro),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F8F9FA')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, -1), (-1, -1), verde),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
    ]
    
    # Aplicar negrita a filas específicas
    for fila_idx in filas_negrita:
        estilo_tabla.append(('FONTNAME', (0, fila_idx), (-1, fila_idx), 'Helvetica-Bold'))
    
    desglose_table.setStyle(TableStyle(estilo_tabla))
    
    story.append(desglose_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Información adicional
    if calculo.otrosi_referencia:
        story.append(Paragraph('INFORMACIÓN DE REFERENCIA', subtitulo_estilo))
        ref_data = [
            ['Otro Sí de Referencia:', calculo.otrosi_referencia.numero_otrosi],
            ['Vigencia desde:', calculo.otrosi_referencia.effective_from.strftime('%d/%m/%Y')],
        ]
        ref_table = Table(ref_data, colWidths=[2*inch, 5*inch])
        ref_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), oscuro),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#F8F9FA')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(ref_table)
        story.append(Spacer(1, 0.2*inch))
    
    if calculo.observaciones:
        obs_titulo_estilo = ParagraphStyle(
            'ObsTitulo', parent=texto_estilo, fontSize=10, textColor=oscuro,
            spaceAfter=6, fontName='Helvetica-Bold',
        )
        story.append(Paragraph('Observaciones:', obs_titulo_estilo))
        story.append(Paragraph(calculo.observaciones, texto_estilo))
    
    # Pie de página
    story.append(Spacer(1, 0.3*inch))
    ahora_local = timezone.localtime(timezone.now())
    footer = Paragraph(
        f'Documento generado el {ahora_local.strftime("%d/%m/%Y %H:%M:%S")}',
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    )
    story.append(footer)
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

