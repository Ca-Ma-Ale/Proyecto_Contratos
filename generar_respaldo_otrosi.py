import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Novedades OtroSis'

HEADER_FILL = PatternFill('solid', fgColor='1F3864')
CASE1_FILL  = PatternFill('solid', fgColor='FCE4D6')
CASE2_FILL  = PatternFill('solid', fgColor='FFF2CC')
CASE3_FILL  = PatternFill('solid', fgColor='E2EFDA')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
NORMAL      = Font(size=10)
BOLD        = Font(bold=True, size=10)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT        = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN        = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

COLS   = ['Caso', 'Contrato', 'Num OtroSi', 'Tipo', 'Estado',
          'Vigencia Desde', 'Vigencia Hasta', 'Plazo Actual\n(meses)',
          'Meses Calculados', 'Novedad / Correccion Aplicada']
WIDTHS = [7, 50, 12, 18, 12, 14, 14, 14, 16, 55]

for col, (titulo, ancho) in enumerate(zip(COLS, WIDTHS), 1):
    c = ws.cell(row=1, column=col, value=titulo)
    c.font      = HEADER_FONT
    c.fill      = HEADER_FILL
    c.alignment = CENTER
    c.border    = THIN
    ws.column_dimensions[c.column_letter].width = ancho

ws.row_dimensions[1].height = 32
ws.freeze_panes = 'A2'

def escribir_fila(ws, r, caso, contrato, otrosi, tipo, estado,
                  desde, hasta, plazo_actual, meses_calc, novedad, fill):
    datos = [caso, contrato, otrosi, tipo, estado,
             desde, hasta, plazo_actual, meses_calc, novedad]
    for col, val in enumerate(datos, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.fill      = fill
        c.border    = THIN
        c.font      = NORMAL
        c.alignment = CENTER if col in (1, 3, 4, 5, 6, 7, 8, 9) else LEFT
    ws.row_dimensions[r].height = 16

r = 2

# ── CASO 1: 61 registros ───────────────────────────────────────────────────
NOVEDAD1 = 'nuevo_plazo_meses=0 en tipo que no modifica plazo — corregido a null'
caso1 = [
    ('2025-05-01', 'OS-6', 'AMENDMENT', 'EN_REVISION'),
    ('2024-032-CTO-PUBLICIDAD-EDITORIAL PLANETA', 'OS-4', 'AMENDMENT', 'EN_REVISION'),
    ('2022-020', 'OS-5', 'AMENDMENT', 'APROBADO'),
    ('CTO-BODEGA-VARGAS Y LOPEZ ASOCIADOS LTDA', 'OS-3', 'AMENDMENT', 'APROBADO'),
    ('2023-034-CTO-PUBLICIDAD-KFC', 'OS-3', 'AMENDMENT', 'APROBADO'),
    ('2025-061-CTO-CONCESION-CHILANGO NEVERA', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2024-083-CTO-PUBLICIDAD-BOREALIX', 'OS-3', 'AMENDMENT', 'APROBADO'),
    ('2024-146-CTO-CONCESION-PHONETRONICS', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('CTO-SENTHIA', 'OS-11', 'AMENDMENT', 'APROBADO'),
    ('CTO-SENTHIA', 'OS-10', 'AMENDMENT', 'APROBADO'),
    ('2023-040 CTO-BODEGA-ESTIMULARTE', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('035-2024-CTO-AIRE ACONDICIONADO-COLPENSIONES', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2025-05-01', 'OS-5', 'AMENDMENT', 'APROBADO'),
    ('022-2024-CTO-AIRE ACONDICIONADO-COLPENSIONES', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2023-036 CTO-BODEGA-CHILANGO', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('CTO-BODEGA-VARGAS Y LOPEZ ASOCIADOS LTDA', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('2023-020-CTO-ESTACIONAMIENTO-SMART FIT', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('2024-083-CTO-PUBLICIDAD-BOREALIX', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2024-071-CTO-PUBLICIDAD-PLANET GAME', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2023-009-CTO-CONCESION-PAGA TODO', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2023-034-CTO-PUBLICIDAD-KFC', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2022-020', 'OS-3', 'AMENDMENT', 'APROBADO'),
    ('2023-040 CTO-BODEGA-ESTIMULARTE', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2024-032-CTO-PUBLICIDAD-EDITORIAL PLANETA', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2022-020', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('2023-032 CTO-BODEGA-MR BONO', 'OS-2', 'POLIZAS_UPDATE', 'APROBADO'),
    ('2023-032 CTO-BODEGA-MR BONO', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('CTO-BODEGA-VARGAS Y LOPEZ ASOCIADOS LTDA', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2023-036 CTO-BODEGA-CHILANGO', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2023-020-CTO-ESTACIONAMIENTO-SMART FIT', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('2023-014-CTO-PUBLICIDAD-TAQUILLA LIVE', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('POPSY-CTO-01', 'OS-8', 'AMENDMENT', 'APROBADO'),
    ('2022-020', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('POPSY-CTO-01', 'OS-7', 'CANON_CHANGE', 'APROBADO'),
    ('001-10-CPGRC10327', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('CTO-BODEGA-COLOMBIA TELECOMUNICACIONES MOVISTAR', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('POPSY-CTO-01', 'OS-6', 'CANON_CHANGE', 'APROBADO'),
    ('CTO-SCOTIBANK-01', 'OS-8', 'AMENDMENT', 'APROBADO'),
    ('2025-05-01', 'OS-3', 'AMENDMENT', 'APROBADO'),
    ('CTO-SENTHIA', 'OS-8', 'AMENDMENT', 'APROBADO'),
    ('CTO-BODEGA-CREPES&WAFLESS BANOS Y LOCKERS', 'OS-4', 'AMENDMENT', 'APROBADO'),
    ('POPSY-CTO-01', 'OS-5', 'AMENDMENT', 'APROBADO'),
    ('2025-05-01', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('CTO-KEVINS-ORAFA', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('CTO-SENTHIA', 'OS-7', 'AMENDMENT', 'APROBADO'),
    ('CTO-BODEGA-CREPES&WAFLESS BANOS Y LOCKERS', 'OS-3', 'AMENDMENT', 'APROBADO'),
    ('POPSY-CTO-01', 'OS-4', 'AMENDMENT', 'APROBADO'),
    ('POPSY-CTO-01', 'OS-2', 'CANON_CHANGE', 'APROBADO'),
    ('CTO-SENTHIA', 'OS-4', 'AMENDMENT', 'APROBADO'),
    ('CTO-BODEGA-CREPES&WAFLESS BANOS Y LOCKERS', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('CTO-SCOTIBANK-01', 'OS-6', 'AMENDMENT', 'APROBADO'),
    ('CTO-SENTHIA', 'OS-3', 'AMENDMENT', 'APROBADO'),
    ('POPSY-CTO-01', 'OS-1', 'CANON_CHANGE', 'APROBADO'),
    ('CTO-SENTHIA', 'OS-2', 'AMENDMENT', 'APROBADO'),
    ('CTO-SCOTIBANK-01', 'OS-5', 'AMENDMENT', 'APROBADO'),
    ('CTO-SENTHIA', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('CTO-SCOTIBANK-01', 'OS-4', 'AMENDMENT', 'APROBADO'),
    ('CTO-BODEGA-CREPES&WAFLESS BANOS Y LOCKERS', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('CTO-SCOTIBANK-01', 'OS-3', 'AMENDMENT', 'APROBADO'),
    ('CTO-ANTENA-PHOENIX TOWER', 'OS-1', 'AMENDMENT', 'APROBADO'),
    ('CTO-SCOTIBANK-01', 'OS-2', 'AMENDMENT', 'APROBADO'),
]
for d in caso1:
    escribir_fila(ws, r, 1, d[0], d[1], d[2], d[3],
                  None, None, 0, None, NOVEDAD1, CASE1_FILL)
    r += 1

# ── CASO 2: 10 registros ──────────────────────────────────────────────────
caso2 = [
    ('2023-014-CTO-PUBLICIDAD-TAQUILLA LIVE', 'OS-7', 'PLAZO_EXTENSION', 'APROBADO', '2025-07-01', '2025-12-30', 0, 6, 'PLAZO_EXTENSION con plazo=0 — corregido: 6 meses, fecha_final=2025-12-30'),
    ('2024-071-CTO-PUBLICIDAD-PLANET GAME',   'OS-2', 'PLAZO_EXTENSION', 'APROBADO', '2025-02-12', '2025-08-11', 0, 6, 'PLAZO_EXTENSION con plazo=0 — corregido: 6 meses, fecha_final=2025-08-11'),
    ('2024-068-CTO-CONCESION-DOGGER',         'OS-1', 'PLAZO_EXTENSION', 'APROBADO', '2025-02-01', '2025-07-31', 0, 6, 'PLAZO_EXTENSION con plazo=0 — corregido: 6 meses, fecha_final=2025-07-31'),
    ('2024-032-CTO-PUBLICIDAD-EDITORIAL PLANETA', 'OS-2', 'PLAZO_EXTENSION', 'APROBADO', '2024-11-01', '2025-04-30', 0, 6, 'PLAZO_EXTENSION con plazo=0 — corregido: 6 meses, fecha_final=2025-04-30'),
    ('2023-014-CTO-PUBLICIDAD-TAQUILLA LIVE', 'OS-5', 'PLAZO_EXTENSION', 'APROBADO', '2024-10-01', '2024-12-30', 0, 3, 'PLAZO_EXTENSION con plazo=0 — corregido: 3 meses, fecha_final=2024-12-30'),
    ('2023-014-CTO-PUBLICIDAD-TAQUILLA LIVE', 'OS-4', 'PLAZO_EXTENSION', 'APROBADO', '2024-07-01', '2024-09-30', 0, 3, 'PLAZO_EXTENSION con plazo=0 — corregido: 3 meses, fecha_final=2024-09-30'),
    ('2023-014-CTO-PUBLICIDAD-TAQUILLA LIVE', 'OS-1', 'PLAZO_EXTENSION', 'APROBADO', '2023-06-24', '2023-08-23', 0, 2, 'PLAZO_EXTENSION con plazo=0 — corregido: 2 meses, fecha_final=2023-08-23'),
    ('CTO-SENTHIA',    'OS-6', 'PLAZO_EXTENSION', 'APROBADO', '2019-06-01', '2019-11-30', 0,  6, 'PLAZO_EXTENSION con plazo=0 — corregido: 6 meses, fecha_final=2019-11-30'),
    ('CTO-SENTHIA',    'OS-5', 'PLAZO_EXTENSION', 'APROBADO', '2018-11-01', '2019-05-31', 0,  7, 'PLAZO_EXTENSION con plazo=0 — corregido: 7 meses, fecha_final=2019-05-31'),
    ('CTO-SCOTIBANK-01','OS-1', 'PLAZO_EXTENSION', 'APROBADO', '2013-05-13', '2014-05-12', 0, 12, 'PLAZO_EXTENSION con plazo=0 — corregido: 12 meses, fecha_final=2014-05-12'),
]
for d in caso2:
    escribir_fila(ws, r, 2, d[0], d[1], d[2], d[3],
                  d[4], d[5], d[6], d[7], d[8], CASE2_FILL)
    r += 1

# ── CASO 3: 7 registros ───────────────────────────────────────────────────
caso3 = [
    ('2022-043-CTO-PUBLICIDAD-LA BRASA ROJA',    'OS-6', 'PLAZO_EXTENSION', 'APROBADO', '2025-07-01', '2025-12-31', None,  6, 'fecha_final=null — corregido: fecha=2025-12-31, plazo=6'),
    ('2024-032-CTO-PUBLICIDAD-EDITORIAL PLANETA', 'OS-3', 'PLAZO_EXTENSION', 'APROBADO', '2025-05-01', '2026-04-30',   11, 12, 'fecha_final=null — corregido: fecha=2026-04-30'),
    ('2024-083-CTO-PUBLICIDAD-BOREALIX',          'OS-2', 'PLAZO_EXTENSION', 'APROBADO', '2025-02-12', '2025-08-11',    6,  6, 'fecha_final=null — corregido: fecha=2025-08-11'),
    ('2022-043-CTO-PUBLICIDAD-LA BRASA ROJA',    'OS-5', 'PLAZO_EXTENSION', 'APROBADO', '2025-01-01', '2025-06-30', None,  6, 'fecha_final=null — corregido: fecha=2025-06-30, plazo=6'),
    ('2023-014-CTO-PUBLICIDAD-TAQUILLA LIVE',    'OS-3', 'PLAZO_EXTENSION', 'APROBADO', '2024-01-01', '2024-06-30',    6,  6, 'fecha_final=null — corregido: fecha=2024-06-30'),
    ('2025-05-01',                               'OS-1', 'PLAZO_EXTENSION', 'APROBADO', '2019-07-16', '2020-01-15',    6,  6, 'fecha_final=null — corregido: fecha=2020-01-15'),
    ('2025-109',                                 'OS-1', 'PLAZO_EXTENSION', 'APROBADO', '2017-12-01', '2019-11-30', None, 24, 'fecha_final=null — corregido: fecha=2019-11-30, plazo=24'),
]
for d in caso3:
    escribir_fila(ws, r, 3, d[0], d[1], d[2], d[3],
                  d[4], d[5], d[6], d[7], d[8], CASE3_FILL)
    r += 1

# ── Leyenda ───────────────────────────────────────────────────────────────
r += 1
ws.cell(row=r, column=1, value='LEYENDA').font = BOLD
r += 1
leyenda = [
    ('Caso 1 (naranja)', 'AMENDMENT / CANON_CHANGE / POLIZAS_UPDATE con nuevo_plazo_meses=0 — corregido a null'),
    ('Caso 2 (amarillo)', 'PLAZO_EXTENSION con nuevo_plazo_meses=0 — corregido con meses calculados y nueva_fecha_final_actualizada'),
    ('Caso 3 (verde)',    'PLAZO_EXTENSION con nueva_fecha_final_actualizada=null — corregido con effective_to'),
]
for texto, desc in leyenda:
    ws.cell(row=r, column=1, value=texto).font = BOLD
    ws.cell(row=r, column=2, value=desc).font  = NORMAL
    r += 1

ruta = 'respaldo_novedades_otrosi_%s.xlsx' % date.today().strftime('%Y%m%d')
wb.save(ruta)
print('Excel generado: %s  (%d registros)' % (ruta, 78))
